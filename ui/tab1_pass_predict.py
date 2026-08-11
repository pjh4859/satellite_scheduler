import os
import csv
import yaml
from datetime import datetime, timedelta, timezone

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QListWidget, 
                             QDateTimeEdit, QSpinBox, QPushButton, QTableWidget, 
                             QTableWidgetItem, QLabel, QFileDialog, QHeaderView, QMessageBox, 
                             QInputDialog, QDialog, QListWidgetItem, QDialogButtonBox, QCheckBox,
                             QRadioButton, QButtonGroup, QGroupBox, QComboBox)
from PyQt6.QtCore import Qt, QUrl
from PyQt6.QtGui import QColor, QDesktopServices, QFont

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar

from core.scheduler import parse_tle_from_dir, parse_stations_from_dir, calculate_passes
from core.exporter import export_to_csv, export_to_yaml, export_to_excel_with_color
from core.tle_fetcher import search_satellites_from_celestrak, download_tle_by_norad_id
from core.plan_parser import normalize_sat_name, load_plan_file, load_plan_excel, load_plan_csv

from ui.dialog_tle_generator import TleFromSepVectorDialog


class HighVisibilityNavigationToolbar(NavigationToolbar):
    def __init__(self, canvas, parent=None):
        super().__init__(canvas, parent)
        self.setStyleSheet("""
            QToolBar { background-color: #F8F9FA; border: 1px solid #CCCCCC; padding: 4px; spacing: 6px; }
            QToolButton { background-color: #FFFFFF; color: #111111; border: 1px solid #B0BEC5; border-radius: 4px; padding: 4px 8px; font-weight: bold; }
            QToolButton:hover { background-color: #E3F2FD; border: 1px solid #2196F3; }
            QToolButton:checked { background-color: #BBDEFB; border: 1px solid #1976D2; }
            QLabel { color: #111111; font-weight: bold; }
        """)


class GanttChartDialog(QDialog):
    def __init__(self, calculated_passes, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📊 Multi-Satellite Pass Allocation Gantt Timeline")
        self.resize(1150, 680)
        self.passes = calculated_passes
        self.fig = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        plt.style.use('dark_background')
        self.fig, ax = plt.subplots(figsize=(12, 6), facecolor='#1E1E1E')
        ax.set_facecolor('#262626')
        
        stations = sorted(list({p['station'] for p in self.passes}))
        st_y_map = {st: i for i, st in enumerate(stations)}
        
        from core.color_manager import color_manager
        
        for p in self.passes:
            st = p['station']
            sat_raw = p['satellite']
            
            y_pos = st_y_map[st]
            aos_dt = p['aos']
            los_dt = p['los']
            is_selected = p.get('selected', True)
            
            aos_num = mdates.date2num(aos_dt)
            los_num = mdates.date2num(los_dt)
            width = los_num - aos_num
            
            if is_selected:
                hex_color, _ = color_manager.get_colors(sat_raw)
                face_color = f"#{hex_color}"
                edge_color = "#FFFFFF"
                alpha = 0.95
                hatch = None
            else:
                face_color = "#3A3A3A"
                edge_color = "#D32F2F"
                alpha = 0.6
                hatch = "///"

            ax.barh(y_pos, width, left=aos_num, height=0.45, 
                    align='center', color=face_color, edgecolor=edge_color, 
                    alpha=alpha, hatch=hatch, linewidth=1.0)
            
            if is_selected and width > 0.0008:
                mid_num = aos_num + (width / 2.0)
                try:
                    sat_clean = sat_raw.split("(")[0].strip()
                    mid_date = mdates.num2date(mid_num)
                    ax.text(mid_date, y_pos, sat_clean, ha='center', va='center', 
                            fontsize=8, fontweight='bold', color='#111111', clip_on=True)
                except Exception:
                    pass

        ax.set_yticks(range(len(stations)))
        ax.set_yticklabels(stations, fontsize=11, fontweight='bold', color='#FFFFFF')
        ax.xaxis_date()
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M', tz=timezone.utc))
        self.fig.autofmt_xdate()
        
        ax.set_title("Timeline Matrix: Allocated Passes (Solid Pastel) vs Collided / Blocked (Hatched Gray)", 
                     fontsize=12, fontweight='bold', color='#FFFFFF', pad=15)
        ax.set_xlabel("Time (UTC)", fontsize=10, fontweight='bold', color='#DDDDDD')
        ax.grid(True, linestyle=':', alpha=0.35, color='#999999')
        
        if stations:
            ax.set_ylim(-0.6, len(stations) - 0.4)
            
        self.fig.subplots_adjust(left=0.12, right=0.96, top=0.90, bottom=0.18)
        
        canvas = FigureCanvas(self.fig)
        toolbar = HighVisibilityNavigationToolbar(canvas, self)
        layout.addWidget(toolbar)
        layout.addWidget(canvas)
        
        btn_close = QPushButton("Close Timeline Chart")
        btn_close.setStyleSheet("background-color: #333333; color: white; font-weight: bold; padding: 6px;")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

    def closeEvent(self, event):
        if self.fig:
            plt.close(self.fig)
        super().closeEvent(event)

    def accept(self):
        if self.fig:
            plt.close(self.fig)
        super().accept()


class EqualizeRuleDialog(QDialog):
    def __init__(self, all_satellites, current_targets=None, current_min_targets=None, current_max_targets=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Satellite Allocation Rules (Min Guarantee & Max Cap)")
        self.resize(620, 480)
        self.all_satellites = sorted(all_satellites)
        
        if current_targets is None:
            self.current_targets = set(self.all_satellites)
        else:
            self.current_targets = set(current_targets)
            
        if current_min_targets is None:
            self.current_min_targets = {sat: 1 for sat in self.all_satellites}
        else:
            self.current_min_targets = current_min_targets

        if current_max_targets is None:
            self.current_max_targets = {sat: 0 for sat in self.all_satellites} # 0 = No Limit
        else:
            self.current_max_targets = current_max_targets
            
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("<b>Set Individual Min Guarantee & Max Pass Limit per Satellite:</b><br><font color='#555555'>• Min Guarantee: 0 ~ 50 (0 = No Guarantee)<br>• Max Limit: 0 ~ 999 (0 = No Limit / Uncapped)</font>"))
        
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Include", "Satellite Name", "Min Guarantee", "Max Limit (0=No Limit)"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        
        self.table.setRowCount(len(self.all_satellites))
        
        for idx, sat in enumerate(self.all_satellites):
            chk_item = QTableWidgetItem()
            chk_item.setCheckState(Qt.CheckState.Checked if sat in self.current_targets else Qt.CheckState.Unchecked)
            self.table.setItem(idx, 0, chk_item)
            
            sat_item = QTableWidgetItem(f"🛰️  {sat}")
            sat_item.setFlags(sat_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 1, sat_item)
            
            min_spin = QSpinBox()
            min_spin.setRange(0, 50)
            min_spin.setValue(self.current_min_targets.get(sat, 1))
            self.table.setCellWidget(idx, 2, min_spin)

            max_spin = QSpinBox()
            max_spin.setRange(0, 999)
            max_spin.setValue(self.current_max_targets.get(sat, 0))
            self.table.setCellWidget(idx, 3, max_spin)
            
        layout.addWidget(self.table)
        
        btn_ctrl_layout = QHBoxLayout()
        btn_sel_all = QPushButton("☑ Select All")
        btn_sel_all.clicked.connect(lambda: self.set_all_checks(True))
        btn_ctrl_layout.addWidget(btn_sel_all)
        
        btn_unsel_all = QPushButton("☒ Clear All")
        btn_unsel_all.clicked.connect(lambda: self.set_all_checks(False))
        btn_ctrl_layout.addWidget(btn_unsel_all)
        layout.addLayout(btn_ctrl_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def set_all_checks(self, checked):
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for idx in range(self.table.rowCount()):
            item = self.table.item(idx, 0)
            if item: item.setCheckState(state)

    def get_results(self):
        selected_sats = set()
        min_targets = {}
        max_targets = {}
        for idx in range(self.table.rowCount()):
            item_chk = self.table.item(idx, 0)
            sat_name = self.all_satellites[idx]
            min_spin = self.table.cellWidget(idx, 2)
            max_spin = self.table.cellWidget(idx, 3)
            
            if item_chk and item_chk.checkState() == Qt.CheckState.Checked:
                selected_sats.add(sat_name)
                
            if min_spin: min_targets[sat_name] = min_spin.value()
            if max_spin: max_targets[sat_name] = max_spin.value()
                
        return selected_sats, min_targets, max_targets


class PassPredictTab(QWidget):
    def __init__(self, main_app):
        super().__init__()
        self.main_app = main_app
        self.tle_dir = "tle"
        self.stations_dir = "stations"
        self.pass_output_dir = "pass_output"
        self.color_mode = "STATION"
        
        # 할당 대상 위성 목록 및 최소/최대 패스 지정 맵 변수
        self.equalize_target_sats = None
        self.min_pass_targets = {}
        self.max_pass_targets = {}
        
        if not os.path.exists(self.pass_output_dir):
            os.makedirs(self.pass_output_dir)
            
        self.init_ui()
        self.refresh_tle_files()
        self.refresh_stations()

    def init_ui(self):
        layout = QHBoxLayout(self)
        left_panel = QVBoxLayout()
        
        left_panel.addWidget(QLabel("<b>1. Detected TLE Files:</b>"))
        self.tle_file_list = QListWidget()
        self.tle_file_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        left_panel.addWidget(self.tle_file_list)
        
        tle_btn_layout = QHBoxLayout()
        self.btn_refresh_tle = QPushButton("Refresh")
        self.btn_refresh_tle.clicked.connect(self.refresh_tle_files)
        tle_btn_layout.addWidget(self.btn_refresh_tle)
        
        self.btn_open_tle_folder = QPushButton("📂 Open Folder")
        self.btn_open_tle_folder.clicked.connect(lambda: self.open_local_folder(self.tle_dir))
        tle_btn_layout.addWidget(self.btn_open_tle_folder)
        
        self.btn_fetch_tle = QPushButton("🌐 Fetch Online")
        self.btn_fetch_tle.setStyleSheet("font-weight: bold; color: #0D47A1;")
        self.btn_fetch_tle.clicked.connect(self.click_fetch_online_tle)
        tle_btn_layout.addWidget(self.btn_fetch_tle)

        self.btn_gen_vector_tle = QPushButton("🚀 TLE from Sep Vector")
        self.btn_gen_vector_tle.setStyleSheet("font-weight: bold; color: #E65100;")
        self.btn_gen_vector_tle.clicked.connect(self.click_generate_tle_from_vector)
        tle_btn_layout.addWidget(self.btn_gen_vector_tle)
        
        left_panel.addLayout(tle_btn_layout)
        
        left_panel.addWidget(QLabel("<b>2. Detected Ground Stations:</b>"))
        self.gs_list = QListWidget()
        self.gs_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        left_panel.addWidget(self.gs_list)
        
        gs_btn_layout = QHBoxLayout()
        self.btn_refresh_gs = QPushButton("Refresh")
        self.btn_refresh_gs.clicked.connect(self.refresh_stations)
        gs_btn_layout.addWidget(self.btn_refresh_gs)
        
        self.btn_open_gs_folder = QPushButton("📂 Open Folder")
        self.btn_open_gs_folder.clicked.connect(lambda: self.open_local_folder(self.stations_dir))
        gs_btn_layout.addWidget(self.btn_open_gs_folder)
        
        left_panel.addLayout(gs_btn_layout)
        
        left_panel.addWidget(QLabel("<b>3. Time Window (UTC):</b>"))
        now_utc_naive = datetime.now(timezone.utc).replace(tzinfo=None)
        
        left_panel.addWidget(QLabel("Start Time:"))
        self.start_time_edit = QDateTimeEdit(now_utc_naive)
        self.start_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.start_time_edit.setCalendarPopup(True)
        left_panel.addWidget(self.start_time_edit)
        
        left_panel.addWidget(QLabel("End Time:"))
        self.end_time_edit = QDateTimeEdit(now_utc_naive + timedelta(days=1))
        self.end_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.end_time_edit.setCalendarPopup(True)
        left_panel.addWidget(self.end_time_edit)
        
        left_panel.addWidget(QLabel("<b>4. Filters & Scheduling Rules:</b>"))
        el_layout = QHBoxLayout()
        el_layout.addWidget(QLabel("Min El (deg):"))
        self.min_el_spin = QSpinBox()
        self.min_el_spin.setRange(0, 90)
        self.min_el_spin.setValue(5)
        el_layout.addWidget(self.min_el_spin)
        left_panel.addLayout(el_layout)
        
        dur_layout = QHBoxLayout()
        dur_layout.addWidget(QLabel("Min Dur (sec):"))
        self.min_dur_spin = QSpinBox()
        self.min_dur_spin.setRange(0, 3600)
        self.min_dur_spin.setValue(300)
        dur_layout.addWidget(self.min_dur_spin)
        left_panel.addLayout(dur_layout)
        
        pass_no_layout = QHBoxLayout()
        pass_no_layout.addWidget(QLabel("Start Pass No.:"))
        self.start_pass_spin = QSpinBox()
        self.start_pass_spin.setRange(1, 99999)
        self.start_pass_spin.setValue(1)
        pass_no_layout.addWidget(self.start_pass_spin)
        left_panel.addLayout(pass_no_layout)
        
        self.chk_equalize_sat = QCheckBox("Equalize Sat Allocation (Fairness)")
        self.chk_equalize_sat.setChecked(True)
        self.chk_equalize_sat.setStyleSheet("font-weight: bold; color: #1B5E20;")
        left_panel.addWidget(self.chk_equalize_sat)

        self.btn_open_equalize_dialog = QPushButton("⚙️ Set Allocation Target & Rules")
        self.btn_open_equalize_dialog.setStyleSheet("background-color: #0288D1; color: white; font-weight: bold; margin-top: 2px; margin-bottom: 2px;")
        self.btn_open_equalize_dialog.clicked.connect(self.click_open_equalize_dialog)
        left_panel.addWidget(self.btn_open_equalize_dialog)
        
        self.lbl_logic_desc = QLabel(
            "<i>ℹ️ Swarm/Multi-sat fair distribution algorithm based on pass count & duration.</i>"
        )
        self.lbl_logic_desc.setWordWrap(True)
        self.lbl_logic_desc.setStyleSheet("color: #555555; font-size: 11px; margin-bottom: 5px; margin-left: 15px;")
        left_panel.addWidget(self.lbl_logic_desc)
        
        self.btn_calculate = QPushButton("Calculate Pass Schedule")
        self.btn_calculate.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
        self.btn_calculate.clicked.connect(self.run_scheduling)
        left_panel.addWidget(self.btn_calculate)
        
        layout.addLayout(left_panel, stretch=1)
        
        # ----------------------------------------------------------------------
        # 우측 패널 (스케쥴 테이블 / 시각화 / 외부 파일 로더)
        # ----------------------------------------------------------------------
        right_panel = QVBoxLayout()
        select_all_layout = QHBoxLayout()
        select_all_layout.addWidget(QLabel("<b>Pass Prediction Timeline Matrix:</b>"))

        # 💡 [신규 확장] 외부 스케쥴 파일 로드 엔진 선택 콤보박스 및 로드 버튼
        select_all_layout.addSpacing(10)
        self.combo_import_engine = QComboBox()
        self.combo_import_engine.addItems([
            "Auto Engine (Standard ➔ DRM 시 xlwings 자동 전환)",
            "Standard Engine (openpyxl / CSV)",
            "DRM Bypass Engine (xlwings)"
        ])
        self.combo_import_engine.setToolTip("사내 DRM(문서 보안)이 걸린 Excel 파일은 DRM Bypass 또는 Auto 모드로 읽습니다.")
        select_all_layout.addWidget(self.combo_import_engine)

        self.btn_import_schedule = QPushButton("📂 Import Schedule File (.xlsx / .csv / .yaml)")
        self.btn_import_schedule.setStyleSheet("background-color: #0288D1; color: white; font-weight: bold; padding: 4px 8px;")
        self.btn_import_schedule.clicked.connect(self.click_import_external_schedule)
        select_all_layout.addWidget(self.btn_import_schedule)
        
        self.btn_open_pass_out = QPushButton("📂 Open Folder")
        self.btn_open_pass_out.clicked.connect(lambda: self.open_local_folder(self.pass_output_dir))
        select_all_layout.addWidget(self.btn_open_pass_out)
        
        select_all_layout.addStretch()

        self.chk_highlight_conflict = QCheckBox("Highlight Conflicts (⚠️)")
        self.chk_highlight_conflict.setChecked(True)
        self.chk_highlight_conflict.setStyleSheet("font-weight: bold; color: #D32F2F;")
        self.chk_highlight_conflict.toggled.connect(self.populate_table)
        select_all_layout.addWidget(self.chk_highlight_conflict)

        # Color Mode 스위치 그룹
        group_color = QGroupBox("Color Mode")
        layout_color = QHBoxLayout(group_color)
        
        self.radio_color_station = QRadioButton("By Station")
        self.radio_color_station.setChecked(True)
        self.radio_color_station.toggled.connect(self.on_color_mode_changed)
        layout_color.addWidget(self.radio_color_station)

        self.radio_color_sat = QRadioButton("By Satellite")
        self.radio_color_sat.toggled.connect(self.on_color_mode_changed)
        layout_color.addWidget(self.radio_color_sat)

        select_all_layout.addWidget(group_color)
        
        self.btn_select_all = QPushButton("☑ Check All")
        self.btn_select_all.setFixedWidth(100)
        self.btn_select_all.clicked.connect(lambda: self.set_all_checkboxes(True))
        select_all_layout.addWidget(self.btn_select_all)
        
        self.btn_unselect_all = QPushButton("☒ Unselect All")
        self.btn_unselect_all.setFixedWidth(100)
        self.btn_unselect_all.clicked.connect(lambda: self.set_all_checkboxes(False))
        select_all_layout.addWidget(self.btn_unselect_all)
        right_panel.addLayout(select_all_layout)
        
        self.lbl_summary_bar = QLabel("📊 Satellite Pass Distribution Summary: Run calculation or import schedule file.")
        self.lbl_summary_bar.setStyleSheet("background-color: #F1F8E9; border: 1px solid #C8E6C9; padding: 6px; font-weight: bold; color: #2E7D32;")
        right_panel.addWidget(self.lbl_summary_bar)
        
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Select", "Ground Station", "Satellite", "Pass No. (Orbit)", "AOS (UTC)", "LOS (UTC)", "Duration (s)", "Max El (deg)", "Status"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setDefaultSectionSize(140)
        
        self.table.itemChanged.connect(self.handle_table_lock)
        right_panel.addWidget(self.table)
        
        btn_layout = QHBoxLayout()
        
        self.btn_gantt_chart = QPushButton("📊 View Gantt Timeline Chart")
        self.btn_gantt_chart.setStyleSheet("background-color: #6A1B9A; color: white; font-weight: bold;")
        self.btn_gantt_chart.clicked.connect(self.click_view_gantt_chart)
        btn_layout.addWidget(self.btn_gantt_chart)
        
        self.btn_csv = QPushButton("Export Selected to CSV")
        self.btn_csv.clicked.connect(self.click_export_csv)
        btn_layout.addWidget(self.btn_csv)
        
        self.btn_excel = QPushButton("🎨 Export Selected to Excel (.xlsx)")
        self.btn_excel.setStyleSheet("font-weight: bold; color: #1E7145;")
        self.btn_excel.clicked.connect(self.click_export_excel)
        btn_layout.addWidget(self.btn_excel)
        
        self.btn_yaml = QPushButton("Export Selected to YAML")
        self.btn_yaml.clicked.connect(self.click_export_yaml)
        btn_layout.addWidget(self.btn_yaml)
        
        right_panel.addLayout(btn_layout)
        layout.addLayout(right_panel, stretch=3)

    # --------------------------------------------------------------------------
    # 💡 [신규/수정] 외부 생성/Export 스케쥴 파일 전용 역로딩 파이프라인
    # --------------------------------------------------------------------------
    def click_import_external_schedule(self):
        """외부 Export 스케쥴 파일 (.xlsx, .csv, .yaml)을 원본 서식대로 직접 읽어 테이블에 바인딩"""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Schedule File", self.pass_output_dir, 
            "Supported Schedule Files (*.xlsx *.xls *.csv *.yaml *.yml)"
        )
        if not path:
            return

        engine_idx = self.combo_import_engine.currentIndex()
        engine_map = {0: "auto", 1: "standard", 2: "xlwings"}
        selected_engine = engine_map[engine_idx]

        try:
            parsed_passes = self.parse_external_schedule_file(path, engine=selected_engine)
            if not parsed_passes:
                QMessageBox.warning(self, "Warning", "The selected schedule file contains no valid pass data.")
                return

            self.main_app.calculated_passes = parsed_passes
            self.populate_table()

            filename = os.path.basename(path)
            QMessageBox.information(
                self, "Import Complete", 
                f"Successfully loaded {len(parsed_passes)} pass records from:\n'{filename}'"
            )
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to load schedule file:\n{str(e)}")

    def parse_external_schedule_file(self, file_path, engine="auto"):
        """Tab 2 변환 로직을 우회하여 패스 스케쥴 원본 데이터 추출"""
        ext = os.path.splitext(file_path)[1].lower()
        passes = []

        # 1. YAML 파일 파싱
        if ext in [".yaml", ".yml"]:
            with open(file_path, "r", encoding="utf-8") as f:
                content = yaml.safe_load(f) or {}
            raw_list = content if isinstance(content, list) else content.get("predicted_passes", content.get("schedule", content.get("passes", [])))
            for item in raw_list:
                if isinstance(item, dict):
                    passes.append(self.convert_dict_to_pass_item(item))
            return passes

        # 2. Excel (.xlsx, .xls) 파싱 (독립 엑셀 로더 적용)
        if ext in [".xlsx", ".xls"]:
            raw_dicts = self._read_raw_excel(file_path, engine=engine)
            for row in raw_dicts:
                passes.append(self.convert_dict_to_pass_item(row))
            return passes

        # 3. CSV 파싱 (YAML 형식 CSV vs 일반 CSV 감지)
        if ext == ".csv":
            # 3-1. YAML 텍스트 구조로 저장된 CSV인 경우 1차 시도
            try:
                with open(file_path, "r", encoding="utf-8-sig") as f:
                    content = yaml.safe_load(f) or {}
                if isinstance(content, dict) and "predicted_passes" in content:
                    raw_list = content.get("predicted_passes", [])
                    for item in raw_list:
                        if isinstance(item, dict):
                            passes.append(self.convert_dict_to_pass_item(item))
                    return passes
            except Exception:
                pass

            # 3-2. 일반 표 형식 CSV 파싱
            raw_dicts = self._read_raw_csv(file_path)
            for row in raw_dicts:
                passes.append(self.convert_dict_to_pass_item(row))
            return passes

        return passes

    def _read_raw_excel(self, file_path, engine="auto"):
        """Tab 2 변환 없이 순수 Excel 딕셔너리 리스트 반환"""
        rows = []
        if engine == "standard":
            rows = self._read_excel_openpyxl_raw(file_path)
        elif engine == "xlwings":
            rows = self._read_excel_xlwings_raw(file_path)
        else: # auto
            try:
                rows = self._read_excel_openpyxl_raw(file_path)
            except Exception:
                rows = self._read_excel_xlwings_raw(file_path)

        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        dict_rows = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            dict_rows.append(row_dict)
        return dict_rows

    def _read_excel_openpyxl_raw(self, file_path):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

    def _read_excel_xlwings_raw(self, file_path):
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        try:
            wb = app.books.open(file_path)
            sheet = wb.sheets[0]
            raw_data = sheet.used_range.value
            wb.close()
            if raw_data and not isinstance(raw_data[0], list):
                raw_data = [raw_data]
            return raw_data or []
        finally:
            app.quit()

    def _read_raw_csv(self, file_path):
        """Tab 2 변환 없이 순수 CSV DictReader 반환"""
        encodings = ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except Exception:
                continue
        return []

    def convert_dict_to_pass_item(self, d):
        """YAML / CSV / Excel 원본 딕셔너리 키 유연 정규화 탐색"""
        if not isinstance(d, dict):
            d = {}

        # 1. 공백, 언더바, 괄호, 대소문자 제거 정규화 맵
        norm_d = {}
        for k, v in d.items():
            if k is not None:
                norm_k = str(k).lower().replace("_", "").replace(" ", "").replace("(", "").replace(")", "").replace("-", "").replace(".", "")
                norm_d[norm_k] = v

        def get_val(candidate_keys, default_val=""):
            for ck in candidate_keys:
                norm_ck = ck.lower().replace("_", "").replace(" ", "").replace("(", "").replace(")", "").replace("-", "").replace(".", "")
                if norm_ck in norm_d:
                    val = norm_d[norm_ck]
                    if val is not None and str(val).strip() != "":
                        return val
            return default_val

        # 2. 패스 스케쥴 항목 정밀 추출
        station = get_val(["station", "groundstation", "gs", "main"], "GS")
        sat = get_val(["satellite", "satid", "sat"], "SAT")
        pass_no_str = str(get_val(["passno", "passnum", "sequenceid"], 1))
        aos_str = str(get_val(["aos", "aosutc", "sub"], ""))
        los_str = str(get_val(["los", "losutc", "remark"], ""))
        dur_str = str(get_val(["durationsec", "duration", "durations", "mindur"], 0))
        max_el_str = str(get_val(["maxelevation", "maxel", "maxeldeg", "maxelevationdeg", "minel"], 0))
        status_str = str(get_val(["status", "reqcap"], "Normal"))

        return self.convert_values_to_pass_item(
            station, sat, pass_no_str, aos_str, los_str, dur_str, max_el_str, status_str
        )

    def parse_dt_string(self, dt_str):
        """다양한 형태의 문자열/datetime 객체를 naive datetime으로 안전 변환"""
        if isinstance(dt_str, datetime):
            return dt_str.replace(tzinfo=None)
            
        if not dt_str:
            return datetime.now(timezone.utc).replace(tzinfo=None)
        
        dt_str = str(dt_str).strip()
        formats = [
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ",
            "%Y/%m/%d %H:%M:%S"
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(dt_str, fmt)
                return dt
            except ValueError:
                continue
        return datetime.now(timezone.utc).replace(tzinfo=None)
    

    def convert_values_to_pass_item(self, station, sat, pass_no_str, aos_str, los_str, dur_str, max_el_str, status_str):
        """문자열 패스 파라미터를 내부 datetime/float 인스턴스로 변환"""
        # Pass 번호 숫자 추출 ("Pass 1" -> 1)
        digits = [s for s in str(pass_no_str) if s.isdigit()]
        p_no = int("".join(digits)) if digits else 1

        # 날짜시간 파싱 (여러 포맷 대응)
        aos_dt = self.parse_dt_string(aos_str)
        los_dt = self.parse_dt_string(los_str)

        try: float_dur = float(dur_str)
        except ValueError: float_dur = 0.0

        try: float_el = float(max_el_str)
        except ValueError: float_el = 0.0

        return {
            'station': str(station).strip(),
            'satellite': str(sat).strip(),
            'pass_no': p_no,
            'aos': aos_dt,
            'los': los_dt,
            'duration': float_dur,
            'max_el': float_el,
            'status': str(status_str).replace("⚠️", "").strip(),
            'selected': True,
            'conflict_group': None
        }
    

    # --------------------------------------------------------------------------
    # 기존 이벤트 및 컨트롤 메서드 유지
    # --------------------------------------------------------------------------
    def click_open_equalize_dialog(self):
        selected_files = [item.text() for item in self.tle_file_list.selectedItems()]
        tle_data = parse_tle_from_dir(self.tle_dir, selected_files)
        
        if not tle_data:
            QMessageBox.warning(self, "Warning", "Please select at least one TLE file first.")
            return
            
        all_sats = list(tle_data.keys())
        
        dialog = EqualizeRuleDialog(
            all_satellites=all_sats,
            current_targets=self.equalize_target_sats,
            current_min_targets=self.min_pass_targets,
            current_max_targets=self.max_pass_targets,
            parent=self
        )
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.equalize_target_sats, self.min_pass_targets, self.max_pass_targets = dialog.get_results()
            
            target_info_tokens = []
            for sat in sorted(self.equalize_target_sats):
                mn = self.min_pass_targets.get(sat, 1)
                mx = self.max_pass_targets.get(sat, 0)
                mx_str = "No Limit" if mx == 0 else f"{mx} max"
                target_info_tokens.append(f"• {sat}: Min {mn} / Max ({mx_str})")
                
            info_msg = "Updated Allocation Rules:\n" + "\n".join(target_info_tokens) if target_info_tokens else "No satellites targeted."
            QMessageBox.information(self, "Rules Updated", info_msg)

    def on_color_mode_changed(self):
        if self.radio_color_sat.isChecked():
            self.color_mode = "SATELLITE"
        else:
            self.color_mode = "STATION"
            
        if getattr(self.main_app, 'calculated_passes', None):
            self.populate_table()

    def click_generate_tle_from_vector(self):
        dialog = TleFromSepVectorDialog(tle_dir=self.tle_dir, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_tle_files()

    def open_local_folder(self, folder_path):
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        abs_path = os.path.abspath(folder_path)
        QDesktopServices.openUrl(QUrl.fromLocalFile(abs_path))

    def refresh_tle_files(self):
        self.tle_file_list.clear()
        parse_tle_from_dir(self.tle_dir)
        if os.path.exists(self.tle_dir):
            for filename in os.listdir(self.tle_dir):
                if filename.endswith(".tle") or filename.endswith(".txt"):
                    self.tle_file_list.addItem(filename)
        for i in range(self.tle_file_list.count()):
            self.tle_file_list.item(i).setSelected(True)

    def click_fetch_online_tle(self):
        query, ok = QInputDialog.getText(
            self, "CelesTrak Satellite Search", "Enter Satellite Keyword or NORAD CatNR (e.g. NEONSAT, ISS, 39634):"
        )
        if not ok or not query.strip(): return

        success, sat_list, msg = search_satellites_from_celestrak(query)
        if not success or not sat_list:
            QMessageBox.warning(self, "Search Failed", f"Search results for '{query}':\n{msg}\n\n(No changes were made.)")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Select Satellites to Fetch ({len(sat_list)} found)")
        dialog.resize(500, 360)
        
        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(QLabel(f"<b>Search Results for '{query}':</b><br>(Use Ctrl / Shift or 'Select All' to download multiple satellites)"))
        
        list_widget = QListWidget()
        list_widget.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        
        for sat in sat_list:
            display_text = f"🛰️  {sat['sat_name']}  |  NORAD ID: {sat['norad_id']}  ({sat['int_designator']})"
            item = QListWidgetItem(display_text)
            item.setData(Qt.ItemDataRole.UserRole, sat)
            list_widget.addItem(item)
            
        dlg_layout.addWidget(list_widget)
        
        select_ctrl_layout = QHBoxLayout()
        btn_sel_all = QPushButton("☑ Select All")
        btn_sel_all.clicked.connect(list_widget.selectAll)
        select_ctrl_layout.addWidget(btn_sel_all)
        
        btn_desel_all = QPushButton("☒ Unselect All")
        btn_desel_all.clicked.connect(list_widget.clearSelection)
        select_ctrl_layout.addWidget(btn_desel_all)
        
        dlg_layout.addLayout(select_ctrl_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            selected_items = list_widget.selectedItems()
            if not selected_items: return
            
            success_count = 0
            failed_sats = []
            
            for item in selected_items:
                sat_data = item.data(Qt.ItemDataRole.UserRole)
                target_norad_id = sat_data['norad_id']
                target_sat_name = sat_data['sat_name']
                
                dl_success, save_path = download_tle_by_norad_id(target_norad_id, target_sat_name, self.tle_dir)
                if dl_success:
                    success_count += 1
                else:
                    failed_sats.append(target_sat_name)
            
            if success_count > 0:
                msg = f"Successfully fetched TLE files for {success_count} satellite(s)."
                if failed_sats:
                    msg += f"\n\nFailed satellites: {', '.join(failed_sats)}"
                QMessageBox.information(self, "Download Complete", msg)
                self.refresh_tle_files()
            else:
                QMessageBox.critical(self, "Download Error", f"Failed to download selected satellites.")

    def refresh_stations(self):
        self.gs_list.clear()
        self.main_app.station_data = parse_stations_from_dir(self.stations_dir)
        for cfg in self.main_app.station_data:
            self.gs_list.addItem(f"{cfg[0]} (Lat: {cfg[1]}, Lon: {cfg[2]}) [Down:{cfg[3]} / Cmd:{cfg[4]}]")
        for i in range(self.gs_list.count()):
            self.gs_list.item(i).setSelected(True)

    def run_scheduling(self):
        selected_files = [item.text() for item in self.tle_file_list.selectedItems()]
        tle_data = parse_tle_from_dir(self.tle_dir, selected_files)
        selected_stations = [self.main_app.station_data[self.gs_list.row(item)] for item in self.gs_list.selectedItems()]
        
        start_dt = self.start_time_edit.dateTime().toPyDateTime()
        end_dt = self.end_time_edit.dateTime().toPyDateTime()
        if start_dt >= end_dt:
            QMessageBox.critical(self, "Time Window Error", "Start Time must be earlier than End Time!")
            return
            
        min_el = self.min_el_spin.value()
        min_dur = self.min_dur_spin.value()
        start_pass_no = self.start_pass_spin.value()
        equalize = self.chk_equalize_sat.isChecked()
        
        if not tle_data or not selected_stations:
            QMessageBox.warning(self, "Warning", "No TLE files or Ground Stations selected.")
            return
            
        self.main_app.calculated_passes = calculate_passes(
            tle_data, selected_stations, start_dt, end_dt, min_el, min_dur, start_pass_no,
            equalize_allocation=equalize,
            equalize_target_sats=self.equalize_target_sats,
            min_pass_targets=self.min_pass_targets,
            max_pass_targets=self.max_pass_targets
        )
        self.populate_table()

    def update_summary_dashboard(self):
        if not self.main_app.calculated_passes:
            self.lbl_summary_bar.setText("📊 Satellite Pass Distribution Summary: No data calculated.")
            return
            
        sat_stats = {}
        for p in self.main_app.calculated_passes:
            sat_clean = p['satellite'].split("(")[0].strip()
            if sat_clean not in sat_stats:
                sat_stats[sat_clean] = {'selected_count': 0, 'total_count': 0, 'total_duration_sec': 0.0}
                
            sat_stats[sat_clean]['total_count'] += 1
            if p.get('selected', False):
                sat_stats[sat_clean]['selected_count'] += 1
                sat_stats[sat_clean]['total_duration_sec'] += float(p.get('duration', 0))
                
        summary_tokens = []
        for sat_name, stats in sorted(sat_stats.items()):
            dur_min = stats['total_duration_sec'] / 60.0
            summary_tokens.append(f"🛰️ <b>{sat_name}</b>: {stats['selected_count']}/{stats['total_count']} passes ({dur_min:.1f} min)")
            
        summary_text = "📊 <b>Pass Allocation Summary:</b> &nbsp;&nbsp;|&nbsp;&nbsp; " + " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(summary_tokens)
        self.lbl_summary_bar.setText(summary_text)

    def populate_table(self):
        if getattr(self.main_app, 'is_populating', False): return
        self.table.setRowCount(0)
        if not self.main_app.calculated_passes: return
            
        self.main_app.is_populating = True
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(len(self.main_app.calculated_passes))
            from core.color_manager import color_manager
            
            show_conflict_highlight = self.chk_highlight_conflict.isChecked()
            
            for row_idx, p in enumerate(self.main_app.calculated_passes):
                chk_item = QTableWidgetItem()
                chk_item.setCheckState(Qt.CheckState.Checked if p.get('selected', True) else Qt.CheckState.Unchecked)
                chk_item.setData(Qt.ItemDataRole.UserRole, (row_idx, p.get('conflict_group', None), p['station']))
                self.table.setItem(row_idx, 0, chk_item)
                
                self.table.setItem(row_idx, 1, QTableWidgetItem(p['station']))
                self.table.setItem(row_idx, 2, QTableWidgetItem(p['satellite']))
                self.table.setItem(row_idx, 3, QTableWidgetItem(f"Pass {p['pass_no']}"))
                
                aos_val = p['aos'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['aos'], datetime) else str(p['aos'])
                los_val = p['los'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['los'], datetime) else str(p['los'])
                
                self.table.setItem(row_idx, 4, QTableWidgetItem(aos_val))
                self.table.setItem(row_idx, 5, QTableWidgetItem(los_val))
                self.table.setItem(row_idx, 6, QTableWidgetItem(str(p['duration'])))
                self.table.setItem(row_idx, 7, QTableWidgetItem(str(p['max_el'])))
                
                status_text = p.get('status', 'Normal')
                status_item = QTableWidgetItem(status_text)
                
                if self.color_mode == "SATELLITE":
                    sat_raw = str(p['satellite'])
                    _, base_color = color_manager.get_colors(sat_raw)
                else:
                    st_raw = str(p['station'])
                    _, base_color = color_manager.get_station_colors(st_raw)

                if "Conflict" in status_text and show_conflict_highlight:
                    status_item.setText(f"⚠️ {status_text}")
                    status_item.setForeground(QColor(180, 0, 0))
                    status_item.setFont(QFont("", -1, QFont.Weight.Bold))
                    row_color = QColor(
                        min(255, base_color.red() + 25),
                        max(0, base_color.green() - 25),
                        max(0, base_color.blue() - 25)
                    )
                else:
                    row_color = base_color

                self.table.setItem(row_idx, 8, status_item)

                for col_idx in range(self.table.columnCount()):
                    cell = self.table.item(row_idx, col_idx)
                    if cell:
                        cell.setBackground(row_color)
                    
            self.update_summary_dashboard()
        finally:
            self.table.blockSignals(False)
            self.main_app.is_populating = False

    def click_view_gantt_chart(self):
        if not self.main_app.calculated_passes:
            QMessageBox.warning(self, "Warning", "Please calculate or import pass schedule first.")
            return
        dialog = GanttChartDialog(self.main_app.calculated_passes, self)
        dialog.exec()

    def handle_table_lock(self, item):
        if self.main_app.is_populating or item.column() != 0: return
        user_data = item.data(Qt.ItemDataRole.UserRole)
        if not user_data: return
            
        current_row, group_id, station_name = user_data
        if group_id is None:
            self.main_app.calculated_passes[current_row]['selected'] = (item.checkState() == Qt.CheckState.Checked)
            self.update_summary_dashboard()
            return
            
        if item.checkState() == Qt.CheckState.Checked:
            self.main_app.is_populating = True
            try:
                for r in range(self.table.rowCount()):
                    if r == current_row:
                        self.main_app.calculated_passes[r]['selected'] = True
                        continue
                    other_item = self.table.item(r, 0)
                    if other_item is not None and other_item.data(Qt.ItemDataRole.UserRole) is not None:
                        o_row, o_group, o_station = other_item.data(Qt.ItemDataRole.UserRole)
                        if o_station == station_name and o_group == group_id:
                            other_item.setCheckState(Qt.CheckState.Unchecked)
                            self.main_app.calculated_passes[r]['selected'] = False
            finally:
                self.main_app.is_populating = False
            self.populate_table()
        else:
            self.main_app.calculated_passes[current_row]['selected'] = False
            self.main_app.is_populating = True
            try: self.populate_table()
            finally: self.main_app.is_populating = False

    def click_export_csv(self):
        if not self.main_app.calculated_passes: return
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV Schedule", self.pass_output_dir, "CSV Files (*.csv)")
        if path: export_to_csv(path, self.main_app.calculated_passes)

    def click_export_yaml(self):
        if not self.main_app.calculated_passes: return
        path, _ = QFileDialog.getSaveFileName(self, "Save YAML Schedule", self.pass_output_dir, "YAML Files (*.yaml)")
        if path: export_to_yaml(path, self.main_app.calculated_passes)

    def set_all_checkboxes(self, check_state):
        if not self.main_app.calculated_passes: return
        self.main_app.is_populating = True
        target_state = Qt.CheckState.Checked if check_state else Qt.CheckState.Unchecked
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item:
                item.setCheckState(target_state)
                self.main_app.calculated_passes[r]['selected'] = check_state
        self.main_app.is_populating = False
        self.populate_table()

    def click_export_excel(self):
        if not self.main_app.calculated_passes: return
        if not any(p['selected'] for p in self.main_app.calculated_passes):
            QMessageBox.warning(self, "Warning", "No passes are selected.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save Colorized Excel Schedule", self.pass_output_dir, "Excel Files (*.xlsx)")
        if path:
            success, msg = export_to_excel_with_color(path, self.main_app.calculated_passes)
            if success:
                QMessageBox.information(self, "Export Success", "Excel file generated successfully!")
            else:
                QMessageBox.critical(self, "Export Error", f"Failed to save Excel file:\n{msg}")