import csv
import yaml
import re
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# ==============================================================================
# [유틸리티] 위성 이름 정규화 함수
# ==============================================================================
def normalize_sat_name(sat_str):
    """
    위성 이름 정규화 함수
    
    [기능 설명]
    - 'NEONSAT-1A(67614)'와 같은 위성 텍스트에서 괄호 및 특수문자/공백을 제거하여
      'NEONSAT1A' 형태로 정규화합니다.
    - color_manager와 연동 시 위성별 일관된 파스텔 색상을 매핑할 수 있도록 지원합니다.
    """
    if not sat_str: 
        return ""
    clean = str(sat_str).split("(")[0].strip()
    clean = re.sub(r'[^A-Za-z0-9]', '', clean).upper()
    return clean


# ==============================================================================
# Tab 1: 패스 예측 결과 내보내기 함수들 (CSV, YAML, Excel)
# ==============================================================================
def export_to_csv(file_path, passes_list):
    """
    Tab 1 패스 예측 스케줄을 CSV 파일로 내보내기
    
    [기능 설명]
    - 선택(selected=True)된 패스 데이터만 추출하여 한글 깨짐이 없는 'utf-8-sig' 인코딩으로 저장합니다.
    - datetime 및 string 형태의 날짜/시간 데이터를 안전하게 포맷팅합니다.
    """
    selected_passes = [p for p in passes_list if p.get('selected', False)]
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Station", "Satellite", "Pass_No", "AOS(UTC)", "LOS(UTC)", "Duration_Sec", "Max_Elevation", "Status"])
        for p in selected_passes:
            aos_str = p['aos'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['aos'], datetime) else str(p['aos'])
            los_str = p['los'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['los'], datetime) else str(p['los'])
            writer.writerow([
                p['station'], p['satellite'], p.get('pass_no', 1),
                aos_str, los_str,
                p['duration'], p['max_el'], p.get('status', 'Normal')
            ])


def export_to_yaml(file_path, passes_list, all_passes=None):
    """
    Tab 1 패스 예측 스케줄을 YAML 파일로 내보내기
    
    [기능 설명]
    - 선택된 패스 목록을 생성 시각 타임스탬프와 함께 정돈된 YAML 규격 문서로 출력합니다.
    - all_passes가 제공될 경우 탈락/충돌 패스를 포함한 전체 후보 풀을 'all_candidate_passes'로 함께 직렬화합니다.
    """
    selected_passes = [p for p in passes_list if p.get('selected', False)]
    formatted_list = []
    for p in selected_passes:
        aos_str = p['aos'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['aos'], datetime) else str(p['aos'])
        los_str = p['los'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['los'], datetime) else str(p['los'])
        
        try: pass_no_val = int(p.get('pass_no', 1))
        except (ValueError, TypeError): pass_no_val = 1
        
        try: dur_val = float(p.get('duration', 0.0))
        except (ValueError, TypeError): dur_val = 0.0
        
        try: el_val = float(p.get('max_el', 0.0))
        except (ValueError, TypeError): el_val = 0.0

        formatted_list.append({
            "station": p['station'], 
            "satellite": p['satellite'], 
            "pass_no": pass_no_val,
            "aos": aos_str, 
            "los": los_str,
            "duration_sec": dur_val, 
            "max_elevation_deg": el_val, 
            "status": p.get('status', 'Normal')
        })
        
    payload = {
        "generation_timestamp": datetime.now(timezone.utc).isoformat(),
        "total_passes_count": len(formatted_list),
        "predicted_passes": formatted_list
    }

    # 전체 패스 풀이 존재할 경우 함께 직렬화 (Tab 3 Cross-Satellite Swap에 활용)
    if all_passes:
        formatted_all = []
        for p in all_passes:
            aos_str = p['aos'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['aos'], datetime) else str(p['aos'])
            los_str = p['los'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['los'], datetime) else str(p['los'])
            try: pass_no_val = int(p.get('pass_no', 1))
            except (ValueError, TypeError): pass_no_val = 1
            try: dur_val = float(p.get('duration', 0.0))
            except (ValueError, TypeError): dur_val = 0.0
            try: el_val = float(p.get('max_el', 0.0))
            except (ValueError, TypeError): el_val = 0.0

            formatted_all.append({
                "station": p['station'], 
                "satellite": p['satellite'], 
                "pass_no": pass_no_val,
                "aos": aos_str, 
                "los": los_str,
                "duration_sec": dur_val, 
                "max_elevation_deg": el_val, 
                "selected": p.get('selected', False),
                "status": p.get('status', 'Normal')
            })
        payload["all_candidate_passes"] = formatted_all

    with open(file_path, "w", encoding="utf-8") as f:
        yaml.dump(payload, f, default_flow_style=False, sort_keys=False, allow_unicode=True)


def export_to_excel_with_color(file_path, passes_list):
    """
    Tab 1 패스 예측 스케줄을 파스텔 색상 포함 Excel 파일로 내보내기
    
    [기능 설명]
    - 지상국별 고유 파스텔 배경색을 각 행(Row)에 적용하여 가시성을 높인 엑셀 문서를 만듭니다.
    - 파일이 이미 열려 있어 발생하는 PermissionError 시 친절한 오류 메시지와 성공 여부(True/False)를 반환합니다.
    """
    try:
        selected_passes = [p for p in passes_list if p.get('selected', False)]
        wb = Workbook()
        ws = wb.active
        ws.title = "Pass Schedule"
        
        headers = ["Station", "Satellite", "Pass_No", "AOS(UTC)", "LOS(UTC)", "Duration_Sec", "Max_Elevation", "Status"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        from core.color_manager import color_manager
        data_font = Font(name="맑은 고딕", size=10)
        
        for row_idx, p in enumerate(selected_passes, start=2):
            aos_str = p['aos'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['aos'], datetime) else str(p['aos'])
            los_str = p['los'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(p['los'], datetime) else str(p['los'])
            
            row_data = [
                p['station'], p['satellite'], f"Pass {p.get('pass_no', 1)}",
                aos_str, los_str,
                p['duration'], p['max_el'], p.get('status', 'Normal')
            ]
            ws.append(row_data)
            
            st_key = p['station'].split("(")[0].strip()
            color_hex, _ = color_manager.get_station_colors(st_key)
            row_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.fill = row_fill
                cell.font = data_font
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(file_path)
        return True, "Success"
    except PermissionError:
        return False, f"Permission Denied: The file '{file_path}' is currently open in Excel. Please close it and try again."
    except Exception as e:
        return False, str(e)


# ==============================================================================
# Tab 2: 미션 제약 조건 내보내기 함수들 (Remark 추가)
# ==============================================================================
def export_constraints_to_csv(file_path, extracted_plan_list, headers_labels):
    """
    Tab 2 제약 조건 목록을 CSV 파일로 내보내기 (Remark 항목 추가)
    """
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers_labels)
        for data in extracted_plan_list:
            writer.writerow([
                data.get("sat_id", data.get("satellite", "")),
                data.get("main", data.get("activity", "")),
                data.get("sub", ""),
                data.get("remark", data.get("remarks", data.get("note", ""))),
                data.get("min_el", ""),
                data.get("req_cap", data.get("required_cap", data.get("x_band_req", ""))),
                data.get("min_dur", data.get("min_duration", data.get("min_pass_contact", ""))),
                data.get("pre_req_main", data.get("pre_activity_sequence_id", "")),
                data.get("sequence_id", data.get("activity_sequence_id", ""))
            ])


def export_constraints_to_excel_color(file_path, extracted_plan_list, headers_labels):
    """
    Tab 2 제약 조건 목록을 위성별 색상이 구분된 Excel 파일로 내보내기 (Remark 항목 추가)
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mission Constraints"
        
        ws.append(headers_labels)
        header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers_labels) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            
        data_font = Font(name="맑은 고딕", size=10)
        from core.color_manager import color_manager
        
        for row_idx, data in enumerate(extracted_plan_list, start=2):
            row_values = [
                data.get("sat_id", data.get("satellite", "")),
                data.get("main", data.get("activity", "")),
                data.get("sub", ""),
                data.get("remark", data.get("remarks", data.get("note", ""))),
                data.get("min_el", ""),
                data.get("req_cap", data.get("required_cap", data.get("x_band_req", ""))),
                data.get("min_dur", data.get("min_duration", data.get("min_pass_contact", ""))),
                data.get("pre_req_main", data.get("pre_activity_sequence_id", "")),
                data.get("sequence_id", data.get("activity_sequence_id", ""))
            ]
            ws.append(row_values)
            
            sat_raw = str(data.get("sat_id", data.get("satellite", ""))).strip()
            sat_clean = normalize_sat_name(sat_raw)
            color_hex, _ = color_manager.get_colors(sat_clean)
            row_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for col_idx in range(1, len(headers_labels) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = data_font
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(file_path)
        return True, "Success"
    except PermissionError:
        return False, f"Permission Denied: The file '{file_path}' is currently open in Excel. Please close it and try again."
    except Exception as e:
        return False, str(e)


# ==============================================================================
# Tab 3: 최종 통합 스케줄 내보내기 함수들 (Remark 항목 추가 & 모드별 색상 구분)
# ==============================================================================
def export_final_schedule_to_csv(file_path, final_data):
    """
    Tab 3 최종 산출 스케줄을 CSV 파일로 내보내기 (Remark 열 추가)
    """
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Station", "Satellite", "Pass_No", "AOS(UTC)", "LOS(UTC)", "Duration_Sec", "Max_Elevation", "Status", "Mission Activity", "Remark"])
        for item in final_data:
            writer.writerow([
                item["station"], item["satellite"], item["pass_no"],
                item["aos"], item["los"], item["duration"], item["max_el"],
                item["status"], item["activity"], item.get("remark", "")
            ])


def export_final_schedule_to_excel(file_path, final_data, color_mode):
    """
    Tab 3 최종 통합 스케줄을 Excel 파일로 내보내기 (Remark 열 추가)
    """
    try:
        from core.color_manager import color_manager
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Final Integrated Schedule"
        
        headers = ["Station", "Satellite", "Pass_No", "AOS(UTC)", "LOS(UTC)", "Duration_Sec", "Max_Elevation", "Status", "Mission Activity", "Remark"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            
        data_font = Font(name="맑은 고딕", size=10)
        
        for row_idx, item in enumerate(final_data, start=2):
            ws.append([
                item["station"], item["satellite"], item["pass_no"],
                item["aos"], item["los"], item["duration"], item["max_el"],
                item["status"], item["activity"], item.get("remark", "")
            ])
            
            if color_mode == "STATION":
                st_key = str(item["station"]).split("(")[0].strip()
                color_hex, _ = color_manager.get_station_colors(st_key)
            else:
                sat_raw = str(item["satellite"])
                sat_clean = normalize_sat_name(sat_raw)
                color_hex, _ = color_manager.get_colors(sat_clean)
                
            row_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = data_font
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(file_path)
        return True, "Success"
    except PermissionError:
        return False, f"Permission Denied: The file '{file_path}' is currently open in Excel. Please close it and try again."
    except Exception as e:
        return False, str(e)