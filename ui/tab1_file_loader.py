import os
import csv
import yaml
from datetime import datetime, timezone


class ExternalScheduleLoader:
    @staticmethod
    def parse_dt_string(dt_str):
        if isinstance(dt_str, datetime):
            return dt_str.replace(tzinfo=None)
            
        if not dt_str:
            return datetime.now(timezone.utc).replace(tzinfo=None)
        
        dt_str = str(dt_str).strip()
        formats = [
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ",
            "%Y/%m/%d %H:%M:%S"
        ]
        for fmt in formats:
            try:
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue
        return datetime.now(timezone.utc).replace(tzinfo=None)

    @classmethod
    def convert_values_to_pass_item(cls, station, sat, pass_no_str, aos_str, los_str, dur_str, max_el_str, status_str):
        digits = [s for s in str(pass_no_str) if s.isdigit()]
        p_no = int("".join(digits)) if digits else 1

        aos_dt = cls.parse_dt_string(aos_str)
        los_dt = cls.parse_dt_string(los_str)

        try: float_dur = float(dur_str)
        except ValueError: float_dur = 0.0

        try: float_el = float(max_el_str)
        except ValueError: float_el = 0.0

        return {
            'station': str(station).strip(),
            'satellite': str(sat).strip(),
            'pass_no': p_no,
            'aos': aos_dt,
            'los': los_dt,
            'duration': float_dur,
            'max_el': float_el,
            'status': str(status_str).replace("⚠️", "").strip(),
            'selected': True,
            'conflict_group': None
        }

    @classmethod
    def convert_dict_to_pass_item(cls, d):
        if not isinstance(d, dict):
            d = {}

        norm_d = {}
        for k, v in d.items():
            if k is not None:
                norm_k = str(k).lower().replace("_", "").replace(" ", "").replace("(", "").replace(")", "").replace("-", "").replace(".", "")
                norm_d[norm_k] = v

        def get_val(candidate_keys, default_val=""):
            for ck in candidate_keys:
                norm_ck = ck.lower().replace("_", "").replace(" ", "").replace("(", "").replace(")", "").replace("-", "").replace(".", "")
                if norm_ck in norm_d:
                    val = norm_d[norm_ck]
                    if val is not None and str(val).strip() != "":
                        return val
            return default_val

        station = get_val(["station", "groundstation", "gs", "main"], "GS")
        sat = get_val(["satellite", "satid", "sat"], "SAT")
        pass_no_str = str(get_val(["passno", "passnum", "sequenceid"], 1))
        aos_str = str(get_val(["aos", "aosutc", "sub"], ""))
        los_str = str(get_val(["los", "losutc", "remark"], ""))
        dur_str = str(get_val(["durationsec", "duration", "durations", "mindur"], 0))
        max_el_str = str(get_val(["maxelevation", "maxel", "maxeldeg", "maxelevationdeg", "minel"], 0))
        status_str = str(get_val(["status", "reqcap"], "Normal"))

        return cls.convert_values_to_pass_item(
            station, sat, pass_no_str, aos_str, los_str, dur_str, max_el_str, status_str
        )

    @classmethod
    def parse_external_schedule_file(cls, file_path, engine="auto"):
        ext = os.path.splitext(file_path)[1].lower()
        passes = []

        if ext in [".yaml", ".yml"]:
            with open(file_path, "r", encoding="utf-8") as f:
                content = yaml.safe_load(f) or {}
            raw_list = content if isinstance(content, list) else content.get("predicted_passes", content.get("schedule", content.get("passes", [])))
            for item in raw_list:
                if isinstance(item, dict):
                    passes.append(cls.convert_dict_to_pass_item(item))
            return passes

        if ext in [".xlsx", ".xls"]:
            raw_dicts = cls._read_raw_excel(file_path, engine=engine)
            for row in raw_dicts:
                passes.append(cls.convert_dict_to_pass_item(row))
            return passes

        if ext == ".csv":
            try:
                with open(file_path, "r", encoding="utf-8-sig") as f:
                    content = yaml.safe_load(f) or {}
                if isinstance(content, dict) and "predicted_passes" in content:
                    raw_list = content.get("predicted_passes", [])
                    for item in raw_list:
                        if isinstance(item, dict):
                            passes.append(cls.convert_dict_to_pass_item(item))
                    return passes
            except Exception:
                pass

            raw_dicts = cls._read_raw_csv(file_path)
            for row in raw_dicts:
                passes.append(cls.convert_dict_to_pass_item(row))
            return passes

        return passes

    @classmethod
    def _read_raw_excel(cls, file_path, engine="auto"):
        rows = []
        if engine == "standard":
            rows = cls._read_excel_openpyxl_raw(file_path)
        elif engine == "xlwings":
            rows = cls._read_excel_xlwings_raw(file_path)
        else:
            try:
                rows = cls._read_excel_openpyxl_raw(file_path)
            except Exception:
                rows = cls._read_excel_xlwings_raw(file_path)

        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        dict_rows = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            dict_rows.append(row_dict)
        return dict_rows

    @staticmethod
    def _read_excel_openpyxl_raw(file_path):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

    @staticmethod
    def _read_excel_xlwings_raw(file_path):
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        try:
            wb = app.books.open(file_path)
            sheet = wb.sheets[0]
            raw_data = sheet.used_range.value
            wb.close()
            if raw_data and not isinstance(raw_data[0], list):
                raw_data = [raw_data]
            return raw_data or []
        finally:
            app.quit()

    @staticmethod
    def _read_raw_csv(file_path):
        encodings = ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except Exception:
                continue
        return []