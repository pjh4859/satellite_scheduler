import os
import csv
import yaml
import re
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# ==============================================================================
# [설정] 미션 플랜(제약 조건) 표준 헤더 및 기본 데이터 정의
# ==============================================================================
PLAN_HEADERS = {
    "sat_id": "Sat_ID",
    "main": "Main",
    "sub": "Sub",
    "remark": "Remark",
    "min_el": "Min_El",
    "req_cap": "Required_Cap",
    "min_dur": "Min_Duration",
    "pre_req_main": "Pre_Req_Main",
    "sequence_id": "Sequence_ID"
}

# 기본 생성용 미션 플랜 시퀀스 샘플 데이터 (20개 항목)
DEFAULT_PLAN_ROWS = [
    ["NEONSAT1", "First Contact", "TM 확인\nPanel 전개\nTLE update", "test1", "10", "CMD", "300", "NONE", "1"],
    ["NEONSAT1", "Health check", "EPS / AOCS / S-band 상태 점검", "test2", "10", "NONE", "180", "First Contact", "2"],
    ["NEONSAT1", "Configuration check", "Thermal / STX / FCS param 확인", "test3", "10", "NONE", "180", "Health check", "3"],
    ["NEONSAT1", "X-band Key update", "X key update", "", "15", "CMD", "120", "Configuration check", "4"],
    ["NEONSAT1", "1차 지상 촬영", "시나리오 등록 / 이미징 / 다운로드", "", "20", "DOWN", "400", "X-band Key update", "5"],
    ["NEONSAT1", "1차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "1차 지상 촬영", "6"],
    ["NEONSAT1", "2차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "1차 Cal Maneuver", "7"],
    ["NEONSAT1", "3차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "2차 Cal Maneuver", "8"],
    ["NEONSAT1", "FCS Parameter Update", "FCP 업데이트 (STS-Gyro up)", "", "10", "CMD", "180", "3차 Cal Maneuver", "9"],
    ["NEONSAT1", "1차 별촬영", "시나리오 등록 / Max logging / AOTM 다운", "", "20", "DOWN", "360", "FCS Parameter Update", "10"],
    ["NEONSAT-1A", "First Contact", "TM 확인 / Panel 전개 / TLE update", "", "10", "CMD", "300", "NONE", "1"],
    ["NEONSAT-1A", "Health check", "EPS / AOCS / S-band 상태 점검", "", "10", "NONE", "180", "First Contact", "2"],
    ["NEONSAT-1A", "Configuration check", "Thermal / STX / FCS param 확인", "", "10", "NONE", "180", "Health check", "3"],
    ["NEONSAT-1A", "X-band Key update", "X key update", "", "15", "CMD", "120", "Configuration check", "4"],
    ["NEONSAT-1A", "1차 지상 촬영", "시나리오 등록 / 이미징 / 다운로드", "", "20", "DOWN", "400", "X-band Key update", "5"],
    ["NEONSAT-1A", "1차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "1차 지상 촬영", "6"],
    ["NEONSAT-1A", "2차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "1차 Cal Maneuver", "7"],
    ["NEONSAT-1A", "3차 Cal Maneuver", "시나리오 등록 / Max logging / AOTM 다운", "", "15", "BOTH", "300", "2차 Cal Maneuver", "8"],
    ["NEONSAT-1A", "FCS Parameter Update", "FCP 업데이트 (STS-Gyro up)", "", "10", "CMD", "180", "3차 Cal Maneuver", "9"],
    ["NEONSAT-1A", "1차 별촬영", "시나리오 등록 / Max logging / AOTM 다운", "", "20", "DOWN", "360", "FCS Parameter Update", "10"]
]


# ==============================================================================
# [유틸리티] 위성 이름 정규화 함수
# ==============================================================================
def normalize_sat_name(sat_str):
    """
    위성 이름 정규화 처리 함수
    """
    if not sat_str:
        return ""
    clean = str(sat_str).split("(")[0].strip()
    clean = re.sub(r'[^A-Za-z0-9]', '', clean).upper()
    return clean


# ==============================================================================
# [파싱] 단일 행(Row) 유연 매핑 및 자료형 안전 변환 함수
# ==============================================================================
def parse_row_flexible(row_dict):
    """
    다양한 양식의 CSV / Excel / YAML 데이터 행(Row)을 표준 딕셔너리로 변환
    """
    if not isinstance(row_dict, dict):
        row_dict = {}

    col_map = {str(k).strip().lower().replace("_", "").replace(" ", ""): v for k, v in row_dict.items() if k is not None}
    
    # 1. Sat_ID
    sat = (row_dict.get('Sat_ID') or row_dict.get('sat_id') or 
           row_dict.get('Satellite') or row_dict.get('satellite') or 
           col_map.get('satid') or col_map.get('satellite') or '')
    
    # 2. Main
    main = (row_dict.get('Main') or row_dict.get('main') or 
            row_dict.get('Activity') or row_dict.get('activity') or 
            col_map.get('mainactivity') or col_map.get('activity') or '')
    
    # 3. Sub
    sub = (row_dict.get('Sub') or row_dict.get('sub') or 
           col_map.get('subactivity') or '')
    
    # 4. Remark
    remark = (row_dict.get('Remark') or row_dict.get('remark') or 
              row_dict.get('Remarks') or row_dict.get('remarks') or 
              row_dict.get('Note') or row_dict.get('note') or 
              col_map.get('remark') or col_map.get('remarks') or col_map.get('note') or '')

    # 5. Min_El
    min_el = (row_dict.get('Min_El') or row_dict.get('min_el') or 
              col_map.get('minel') or 0)
    
    # 6. Required_Cap
    req_cap = (row_dict.get('Required_Cap') or row_dict.get('req_cap') or 
               row_dict.get('required_cap') or row_dict.get('X-Band 여부') or 
               col_map.get('requiredcap') or col_map.get('xband여부') or 'NONE')
    
    # 7. Min_Duration
    min_dur = (row_dict.get('Min_Duration') or row_dict.get('min_dur') or 
               row_dict.get('min_duration') or row_dict.get('최소 pass contact 시간') or 
               col_map.get('minduration') or col_map.get('최소passcontact시간') or 0)
    
    # 8. Pre_Req_Main
    pre_req = (row_dict.get('Pre_Req_Main') or row_dict.get('pre_req_main') or 
               row_dict.get('Pre Activity Sequence ID') or 
               col_map.get('prereqmain') or col_map.get('preactivitysequenceid') or 'NONE')
    
    # 9. Sequence_ID
    seq_id = (row_dict.get('Sequence_ID') or row_dict.get('sequence_id') or 
              row_dict.get('activity_sequence_id') or 
              col_map.get('sequenceid') or col_map.get('activitysequenceid') or 999)

    req_cap_str = str(req_cap).strip().upper() if req_cap is not None else 'NONE'
    if req_cap_str == 'Y':
        req_cap_str = 'DOWN'
    elif req_cap_str == 'N':
        req_cap_str = 'NONE'

    try:
        float_el = float(min_el) if min_el is not None else 0.0
    except (ValueError, TypeError):
        float_el = 0.0
    
    try:
        float_dur = float(min_dur) if min_dur is not None else 0.0
    except (ValueError, TypeError):
        float_dur = 0.0

    try:
        int_seq = int(seq_id) if seq_id is not None else 999
    except (ValueError, TypeError):
        int_seq = 999

    return {
        'sat_id': str(sat).strip() if sat is not None else '',
        'main': str(main).strip() if main is not None else '',
        'sub': str(sub).strip() if sub is not None else '',
        'remark': str(remark).strip() if remark is not None else '',
        'min_el': float_el,
        'req_cap': req_cap_str,
        'min_dur': float_dur,
        'pre_req_main': str(pre_req).strip() if pre_req is not None else 'NONE',
        'sequence_id': int_seq
    }


# ==============================================================================
# [초기화] 기본 미션 플랜 샘플 파일 생성 함수들
# ==============================================================================
def create_default_plan_csv(plans_dir):
    if not os.path.exists(plans_dir):
        os.makedirs(plans_dir)
    csv_path = os.path.join(plans_dir, "default_mission_plan.csv")
    if not os.path.exists(csv_path):
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(list(PLAN_HEADERS.values()))
            for row in DEFAULT_PLAN_ROWS:
                writer.writerow(row)


def create_default_plan_excel(plans_dir):
    if not os.path.exists(plans_dir):
        os.makedirs(plans_dir)
    xlsx_path = os.path.join(plans_dir, "default_mission_plan.xlsx")
    if os.path.exists(xlsx_path): 
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Mission Constraints"
    headers = list(PLAN_HEADERS.values())
    ws.append(headers)
    for row in DEFAULT_PLAN_ROWS:
        ws.append([
            row[0], row[1], row[2], row[3],
            float(row[4]), row[5], float(row[6]), row[7], int(row[8])
        ])
    wb.save(xlsx_path)


# ==============================================================================
# [로더] 포맷별 미션 플랜 로딩 함수들
# ==============================================================================
def load_plan_yaml(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        content = yaml.safe_load(f) or {}
        
    if isinstance(content, list):
        raw_list = content
    elif isinstance(content, dict):
        raw_list = content.get("mission_constraints", content.get("constraints", []))
    else:
        raw_list = []
        
    parsed_rows = [parse_row_flexible(item) for item in raw_list if isinstance(item, dict)]
    return parsed_rows


def load_plan_csv(file_path):
    encodings = ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']
    lines = []
    for enc in encodings:
        try:
            with open(file_path, "r", encoding="utf-8" if enc == 'utf-8' else enc) as f:
                lines = [line for line in f if not line.strip().startswith("#")]
            break
        except Exception:
            continue

    if not lines:
        return []
        
    reader = csv.DictReader(lines)
    parsed_rows = [parse_row_flexible(row) for row in reader]
    return parsed_rows


def _read_excel_openpyxl(file_path):
    """openpyxl 기반 빠른 Excel 읽기"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_excel_xlwings(file_path):
    """xlwings 기반 DRM 우회 Excel 읽기 (MS Excel 필요)"""
    try:
        import xlwings as xw
    except ImportError:
        raise ImportError(
            "DRM 우회 기능을 사용하려면 'xlwings' 라이브러리가 필요합니다.\n"
            "터미널에서 'pip install xlwings' 명령어로 라이브러리를 설치해 주세요."
        )

    app = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        wb = app.books.open(file_path)
        sheet = wb.sheets[0]
        raw_data = sheet.used_range.value
        wb.close()

        if not raw_data:
            return []

        if not isinstance(raw_data[0], list):
            raw_data = [raw_data]

        return raw_data
    except Exception as e:
        raise RuntimeError(f"xlwings 읽기 실패 (MS Excel 설치 및 DRM 로그인 상태 확인 필요): {str(e)}")
    finally:
        if app is not None:
            app.quit()


def load_plan_excel(file_path, engine="auto"):
    """
    Excel (.xlsx, .xls) 파일 전용 로더 (DRM 엔진 선택 지원)
    
    :param file_path: 파일 경로
    :param engine: 'auto' (기본값: openpyxl 시도 후 실패 시 xlwings 자동 전환), 
                   'standard' (openpyxl 전용), 
                   'xlwings' (DRM 우회 전용)
    """
    rows = []
    if engine == "standard":
        rows = _read_excel_openpyxl(file_path)
    elif engine == "xlwings":
        rows = _read_excel_xlwings(file_path)
    else:  # "auto"
        try:
            rows = _read_excel_openpyxl(file_path)
        except Exception as e:
            print(f"[plan_parser] openpyxl failed (DRM suspected): {e}")
            print(f"[plan_parser] Retrying with xlwings DRM bypass engine...")
            rows = _read_excel_xlwings(file_path)

    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed_rows = []

    for row in rows[1:]:
        if not any(row):
            continue
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        parsed_rows.append(parse_row_flexible(row_dict))
    return parsed_rows


def load_plan_file(file_path, engine="auto"):
    """
    통합 미션 플랜 파일 로더 (확장자 자동 판별 및 DRM 엔진 지원)
    """
    if not file_path or not os.path.exists(file_path):
        return []

    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".yaml", ".yml"]:
        return load_plan_yaml(file_path)
    elif ext in [".xlsx", ".xls"]:
        return load_plan_excel(file_path, engine=engine)
    elif ext == ".csv":
        return load_plan_csv(file_path)
    else:
        return load_plan_csv(file_path)


# ==============================================================================
# [저장] 미션 플랜 제약 조건 YAML 저장 함수
# ==============================================================================
def save_plan_to_yaml(dest_path, plan_data_list):
    sanitized_list = [parse_row_flexible(item) if isinstance(item, dict) else item for item in plan_data_list]

    payload = {
        "generation_timestamp": datetime.now(timezone.utc).isoformat(),
        "total_constraints_count": len(sanitized_list),
        "mission_constraints": sanitized_list
    }

    dest_dir = os.path.dirname(dest_path)
    if dest_dir and not os.path.exists(dest_dir):
        os.makedirs(dest_dir)

    with open(dest_path, "w", encoding="utf-8") as f:
        yaml.dump(payload, f, default_flow_style=False, sort_keys=False, allow_unicode=True)